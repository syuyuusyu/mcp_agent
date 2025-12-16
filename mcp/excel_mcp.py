
import openpyxl
import pandas as pd
import os
from langchain_core.tools import tool
from typing import Any
import boto3
from app.utils import load_config_yaml
import io


project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
files_dir = os.path.join(project_root, "files")

config = load_config_yaml("config.yaml")

oss_config = config.get("oss",{})

s3_client = boto3.client(
    's3',
    endpoint_url=oss_config.get("endpoint"),  # MinIO 端点
    aws_access_key_id=oss_config.get("access-key"),
    aws_secret_access_key=oss_config.get("secret-key"),
)

def get_file_stream_from_s3(file_name: str) -> io.BytesIO:
    """从 S3 获取文件并返回字节流"""
    try:
        response = s3_client.get_object(
            Bucket=oss_config.get("bucket-name"),
            Key=f"mcp_file/{file_name}"
        )
        file_content = response['Body'].read()
        return io.BytesIO(file_content)
    except Exception as e:
        raise ValueError(f"Failed to get file from S3: {str(e)}")


@tool("read_sheet_names")
def read_sheet_names(file_name: str) -> list[str]:
    """ 读取 Excel 文件的工作表名称"""
    try:
        file_stream = get_file_stream_from_s3(file_name)
        
        # 用 openpyxl 打开
        workbook = openpyxl.load_workbook(file_stream, data_only=False)
        sheet_names = workbook.sheetnames
        workbook.close()
        return sheet_names
    except Exception as e:
        raise ValueError(f"Failed to read from S3: {str(e)}")

@tool("read_sheet_data")
def read_sheet_data(file_name: str, sheetName: str) -> list[list]:
    """读取 Excel 工作表中的数据。
    Args:
        file_name: Excel 文件名(文件路径由程序本身确定，模型不需要传入路径)
        sheetName: 工作表名称
    Returns:
        二维数组形式的数据
    """
    try:
        file_stream = get_file_stream_from_s3(file_name)
        df = pd.read_excel(file_stream, sheet_name=sheetName)
        return df.values.tolist()
    except Exception as e:
        raise ValueError(f"Failed to read sheet data: {str(e)}")

@tool("read_sheet_formula")
def read_sheet_formula(file_name: str, sheetName: str) -> list[list[str]]:
    """读取 Excel 工作表中的公式。
    Args:
        file_name: Excel 文件名(文件路径由程序本身确定，模型不需要传入路径)
        sheetName: 工作表名称
    Returns:
        二维数组形式的公式
    """
    try:
        file_stream = get_file_stream_from_s3(file_name)
        workbook = openpyxl.load_workbook(file_stream, data_only=False)
        sheet = workbook[sheetName]
        formulas = []
        for row in sheet.iter_rows():
            row_formulas = []
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    row_formulas.append(cell.value)
                else:
                    row_formulas.append(None)
            formulas.append(row_formulas)
        workbook.close()
        return formulas
    except Exception as e:
        raise ValueError(f"Failed to read sheet formula: {str(e)}")

@tool("write_sheet_data")
def write_sheet_data(file_name: str, sheetName: str, data: list[list]) -> bool:
    """写入数据到 Excel 工作表。
    Args:
        file_name: Excel 文件名(文件路径由程序本身确定，模型不需要传入路径)
        sheetName: 工作表名称
        data: 要写入的二维数组数据
    Returns:
        是否写入成功
    """
    try:
        # 尝试从 S3 下载现有文件，如果不存在则创建新的
        try:
            file_stream = get_file_stream_from_s3(file_name)
            workbook = openpyxl.load_workbook(file_stream)
        except:
            workbook = openpyxl.Workbook()
        
        # 创建或获取工作表
        if sheetName not in workbook.sheetnames:
            workbook.create_sheet(sheetName)
        else:
            # 清空现有数据
            sheet = workbook[sheetName]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
        
        sheet = workbook[sheetName]
        
        # 写入数据
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                sheet.cell(row=i+1, column=j+1, value=value)
        
        # 保存到临时流，然后上传到 S3
        output_stream = io.BytesIO()
        workbook.save(output_stream)
        workbook.close()
        
        output_stream.seek(0)
        s3_client.put_object(
            Bucket=oss_config.get("bucket-name"),
            Key=f"mcp_file/{file_name}",
            Body=output_stream.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return True
    except Exception as e:
        raise ValueError(f"Failed to write sheet data: {str(e)}")

@tool("write_sheet_formula")
def write_sheet_formula(file_name: str, sheetName: str, formulas: list[list[str]]) -> bool:
    """写入公式到 Excel 工作表。
    Args:
        file_name: Excel 文件名(文件路径由程序本身确定，模型不需要传入路径)
        sheetName: 工作表名称
        formulas: 要写入的二维数组公式
    Returns:
        是否写入成功
    """
    try:
        # 尝试从 S3 下载现有文件，如果不存在则创建新的
        try:
            file_stream = get_file_stream_from_s3(file_name)
            workbook = openpyxl.load_workbook(file_stream)
        except:
            workbook = openpyxl.Workbook()
        
        # 创建或获取工作表
        if sheetName not in workbook.sheetnames:
            workbook.create_sheet(sheetName)
        
        sheet = workbook[sheetName]
        
        # 写入公式
        for i, row in enumerate(formulas):
            for j, formula in enumerate(row):
                if formula:
                    sheet.cell(row=i+1, column=j+1, value=formula)
        
        # 保存到临时流，然后上传到 S3
        output_stream = io.BytesIO()
        workbook.save(output_stream)
        workbook.close()
        
        output_stream.seek(0)
        s3_client.put_object(
            Bucket=oss_config.get("bucket-name"),
            Key=f"mcp_file/{file_name}",
            Body=output_stream.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return True
    except Exception as e:
        raise ValueError(f"Failed to write sheet formula: {str(e)}")

@tool("create_excel_file")
def create_excel_file(file_name: str, title: list[str], data: list[list[Any]]) -> bool:
    """创建新的 Excel 文件并写入数据。
    Args:
        file_name: 要创建的 Excel 文件名(文件路径由程序本身确定，模型不需要传入路径)
        title: 表头列表，例如 ["姓名", "年龄", "成绩"]
        data: 要写入的数据，二维数组，每个内部数组代表一行数据
    Returns:
        是否创建成功
    """
    try:
        # 创建 DataFrame
        df = pd.DataFrame(data, columns=title)
        
        # 保存到临时流
        output_stream = io.BytesIO()
        with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # 上传到 S3
        output_stream.seek(0)
        s3_client.put_object(
            Bucket=oss_config.get("bucket-name"),
            Key=f"mcp_file/{file_name}",
            Body=output_stream.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return True
    except Exception as e:
        raise ValueError(f"Failed to create Excel file: {str(e)}")
